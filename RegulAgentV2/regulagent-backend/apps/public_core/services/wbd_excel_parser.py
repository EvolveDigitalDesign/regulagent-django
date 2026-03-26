"""
WBD Excel Parser

Parses Well Bore Diagram (WBD) Excel workbooks exported from the WBD generator.
Reads Sheet 2 ("Plug Details") and Sheet 3 ("Well Geometry").
"""

import logging
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Public constants
# ---------------------------------------------------------------------------

VALID_PLUG_TYPES = {"cement", "bridge_plug", "cast_iron_bridge_plug"}
VALID_PLACEMENT_METHODS = {"pump_and_plug", "balanced", "dump_bailer"}

SHEET_PLUG_DETAILS = "Plug Details"
SHEET_WELL_GEOMETRY = "Well Geometry"

# Named ranges produced by the WBD generator
NAMED_RANGE_PLUG_DETAILS = "PlugDetails"
NAMED_RANGE_CASING = "CasingRecord"
NAMED_RANGE_FORMATION = "FormationTops"
NAMED_RANGE_PERFORATIONS = "Perforations"
NAMED_RANGE_TUBING = "Tubing"
NAMED_RANGE_TOOLS = "ToolsEquipment"

# Section header keywords used for fallback scanning (Sheet 3)
SECTION_KEYWORDS = {
    "casing_strings": ["casing record", "casing"],
    "formation_tops": ["formation tops", "formation top"],
    "perforations": ["perforations", "perforation"],
    "tubing": ["tubing"],
    "tools": ["tools/equipment", "tools & equipment", "tools and equipment", "tools"],
}


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------


def _safe_float(val: Any) -> float | None:
    """Convert a cell value to float, returning None on failure."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip().replace(",", "")
        if val == "":
            return None
        try:
            return float(val)
        except ValueError:
            return None
    return None


def _safe_int(val: Any) -> int | None:
    """Convert a cell value to int, returning None on failure."""
    f = _safe_float(val)
    if f is None:
        return None
    return int(f)


def _safe_str(val: Any) -> str | None:
    """Strip whitespace and return lowercase string, or None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _normalize_type(raw: str | None) -> str | None:
    """Normalize a plug type string to a canonical value."""
    if raw is None:
        return None
    cleaned = raw.strip().lower().replace(" ", "_").replace("-", "_")
    # Common aliases
    aliases = {
        "cast_iron_bridge": "cast_iron_bridge_plug",
        "cibp": "cast_iron_bridge_plug",
        "bp": "bridge_plug",
        "bridge": "bridge_plug",
    }
    return aliases.get(cleaned, cleaned)


def _normalize_placement(raw: str | None) -> str | None:
    """Normalize a placement method string to a canonical value."""
    if raw is None:
        return None
    cleaned = raw.strip().lower().replace(" ", "_").replace("-", "_")
    aliases = {
        "pump_plug": "pump_and_plug",
        "pump_&_plug": "pump_and_plug",
        "dump_bail": "dump_bailer",
    }
    return aliases.get(cleaned, cleaned)


def _parse_bool(val: Any) -> bool | None:
    """Parse truthy values from Excel into bool."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    s = str(val).strip().lower()
    if s in ("yes", "true", "1", "y"):
        return True
    if s in ("no", "false", "0", "n"):
        return False
    return None


def _is_blank_row(row: tuple) -> bool:
    """Return True if every cell in the row is empty."""
    return all(c.value is None or str(c.value).strip() == "" for c in row)


def _cell_is_bold(cell: Cell) -> bool:
    """Return True if the cell has bold font."""
    try:
        return bool(cell.font and cell.font.bold)
    except Exception:
        return False


def _row_values(row: tuple) -> list:
    """Extract raw values from a row tuple."""
    return [c.value for c in row]


def _header_map(header_row: tuple) -> dict[str, int]:
    """
    Build a mapping from normalized header text → 0-based column index.
    Handles None cells gracefully.
    """
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        if cell.value is not None:
            key = str(cell.value).strip().lower()
            mapping[key] = idx
    return mapping


# ---------------------------------------------------------------------------
# Named-range helpers
# ---------------------------------------------------------------------------


def _named_range_bounds(wb: openpyxl.Workbook, name: str) -> tuple[str, int, int, int, int] | None:
    """
    Resolve a named range to (sheet_title, min_row, min_col, max_row, max_col).
    Returns None if the name is not defined or cannot be resolved.
    """
    try:
        defined = wb.defined_names.get(name)
        if defined is None:
            return None
        destinations = list(defined.destinations)
        if not destinations:
            return None
        sheet_title, coord = destinations[0]
        ws = wb[sheet_title]
        cell_range = ws[coord]
        # cell_range may be a single cell, a row, or a 2D tuple
        if isinstance(cell_range, Cell):
            return (sheet_title, cell_range.row, cell_range.column, cell_range.row, cell_range.column)
        if isinstance(cell_range, tuple):
            if isinstance(cell_range[0], Cell):
                # Single row
                rows = [cell_range]
            else:
                rows = cell_range
            min_r = rows[0][0].row
            max_r = rows[-1][0].row
            min_c = rows[0][0].column
            max_c = rows[0][-1].column
            return (sheet_title, min_r, min_c, max_r, max_c)
        return None
    except Exception as exc:
        logger.debug("Could not resolve named range %s: %s", name, exc)
        return None


def _iter_named_range(wb: openpyxl.Workbook, name: str):
    """
    Yield rows from a named range. Each row is a tuple of Cell objects.
    Returns an empty iterator if the range cannot be resolved.
    """
    bounds = _named_range_bounds(wb, name)
    if bounds is None:
        return
    sheet_title, min_r, min_c, max_r, max_c = bounds
    ws = wb[sheet_title]
    yield from ws.iter_rows(min_row=min_r, min_col=min_c, max_row=max_r, max_col=max_c)


# ---------------------------------------------------------------------------
# Sheet 2: Plug Details
# ---------------------------------------------------------------------------


def _find_plug_header_row(ws: Worksheet) -> int | None:
    """
    Scan Sheet 2 for the header row containing 'plug #', 'type', 'top (ft)'.
    Returns the 1-based row index or None.
    """
    for row in ws.iter_rows():
        values = [str(c.value).strip().lower() if c.value is not None else "" for c in row]
        if "plug #" in values and "type" in values and "top (ft)" in values:
            return row[0].row
    return None


def _parse_plug_details(ws: Worksheet, wb: openpyxl.Workbook, warnings: list[str]) -> list[dict]:
    """Parse Sheet 2 into a list of plug dicts."""
    plugs: list[dict] = []

    # --- Attempt named range first ---
    named_rows = list(_iter_named_range(wb, NAMED_RANGE_PLUG_DETAILS))
    if named_rows:
        # First row of named range should be the header
        header = _header_map(named_rows[0])
        data_rows = named_rows[1:]
    else:
        # Fallback: scan for header row
        header_row_num = _find_plug_header_row(ws)
        if header_row_num is None:
            warnings.append("Plug Details: Could not locate header row — sheet may be empty or malformed")
            return plugs
        all_rows = list(ws.iter_rows())
        header = _header_map(all_rows[header_row_num - 1])
        data_rows = all_rows[header_row_num:]

    # Build a flexible column index resolver
    col = _build_plug_col_map(header, warnings)

    for row in data_rows:
        if _is_blank_row(row):
            continue
        vals = _row_values(row)

        def get(key: str, default=None):
            idx = col.get(key)
            if idx is None or idx >= len(vals):
                return default
            return vals[idx]

        plug_num = _safe_int(get("plug_number"))
        raw_type = _safe_str(get("type"))

        if plug_num is None:
            if raw_type is not None:
                warnings.append(
                    f"Plug Details row {row[0].row}: missing Plug # — skipping row"
                )
            continue
        if raw_type is None:
            warnings.append(f"Plug Details row {row[0].row}: missing Type for plug #{plug_num} — skipping row")
            continue

        top = _safe_float(get("top_ft"))
        bottom = _safe_float(get("bottom_ft"))
        if top is None or bottom is None:
            warnings.append(
                f"Plug Details row {row[0].row}: plug #{plug_num} missing top/bottom depth — skipping row"
            )
            continue

        norm_type = _normalize_type(raw_type)
        if norm_type not in VALID_PLUG_TYPES:
            warnings.append(
                f"Plug Details row {row[0].row}: plug #{plug_num} has unknown type '{raw_type}' "
                f"(expected one of {', '.join(sorted(VALID_PLUG_TYPES))})"
            )

        raw_placement = _safe_str(get("placement_method"))
        norm_placement = _normalize_placement(raw_placement) if raw_placement else None
        if norm_placement is not None and norm_placement not in VALID_PLACEMENT_METHODS:
            warnings.append(
                f"Plug Details row {row[0].row}: plug #{plug_num} has unknown placement method '{raw_placement}'"
            )

        plug = {
            "plug_number": plug_num,
            "type": norm_type,
            "top_ft": top,
            "bottom_ft": bottom,
            "sacks": _safe_float(get("sacks")),
            "cement_class": _safe_str(get("cement_class")),
            "tagged_depth_ft": _safe_float(get("tagged_depth_ft")),
            "placement_method": norm_placement,
            "woc_hours": _safe_float(get("woc_hours")),
            "woc_tagged": _parse_bool(get("woc_tagged")),
        }
        plugs.append(plug)

    return plugs


def _build_plug_col_map(header: dict[str, int], warnings: list[str]) -> dict[str, int]:
    """
    Map logical field names to column indices using flexible header matching.
    """
    patterns: dict[str, list[str]] = {
        "plug_number": ["plug #", "plug#", "plug number", "plug no", "plug no."],
        "type": ["type"],
        "top_ft": ["top (ft)", "top(ft)", "top ft", "top"],
        "bottom_ft": ["bottom (ft)", "bottom(ft)", "bottom ft", "bottom"],
        "sacks": ["sacks", "sacks of cement", "# sacks"],
        "cement_class": ["cement class", "class"],
        "tagged_depth_ft": ["tagged depth (ft)", "tagged depth(ft)", "tagged depth ft", "tagged depth"],
        "placement_method": ["placement method", "placement"],
        "woc_hours": ["woc hours", "woc (hours)", "woc hrs"],
        "woc_tagged": ["woc tagged", "woc_tagged"],
    }
    result: dict[str, int] = {}
    for field, candidates in patterns.items():
        for candidate in candidates:
            if candidate in header:
                result[field] = header[candidate]
                break
    return result


# ---------------------------------------------------------------------------
# Sheet 3: Well Geometry — sub-table detection
# ---------------------------------------------------------------------------


def _section_from_row(row: tuple) -> str | None:
    """
    Determine if a row is a section header for one of the well geometry sub-tables.
    Returns the section key ("casing_strings", "formation_tops", etc.) or None.
    """
    combined_text = " ".join(
        str(c.value).strip().lower() for c in row if c.value is not None
    )
    if not combined_text:
        return None
    for section_key, keywords in SECTION_KEYWORDS.items():
        for kw in keywords:
            if kw in combined_text:
                return section_key
    return None


def _is_section_header(row: tuple) -> bool:
    """Return True if any cell in the row is bold (section header indicator)."""
    return any(_cell_is_bold(c) for c in row if c.value is not None)


def _parse_well_geometry(ws: Worksheet, wb: openpyxl.Workbook, warnings: list[str]) -> dict:
    """Parse Sheet 3 into well geometry sub-tables."""
    geometry: dict = {
        "casing_strings": [],
        "formation_tops": [],
        "perforations": [],
        "tubing": [],
        "tools": [],
    }

    # Try named ranges first
    _try_named_ranges(wb, geometry, warnings)

    # If any section is still empty, fall back to row scanning
    all_empty = all(len(geometry[k]) == 0 for k in geometry)
    if all_empty or not _any_named_ranges_exist(wb):
        _scan_geometry_rows(ws, geometry, warnings)

    return geometry


def _any_named_ranges_exist(wb: openpyxl.Workbook) -> bool:
    """Return True if at least one of the well geometry named ranges exists."""
    for name in [NAMED_RANGE_CASING, NAMED_RANGE_FORMATION, NAMED_RANGE_PERFORATIONS,
                 NAMED_RANGE_TUBING, NAMED_RANGE_TOOLS]:
        if _named_range_bounds(wb, name) is not None:
            return True
    return False


def _try_named_ranges(wb: openpyxl.Workbook, geometry: dict, warnings: list[str]) -> None:
    """Attempt to populate geometry from named ranges."""
    # Casing
    casing_rows = list(_iter_named_range(wb, NAMED_RANGE_CASING))
    if casing_rows:
        header = _header_map(casing_rows[0])
        geometry["casing_strings"] = _parse_casing_rows(casing_rows[1:], header, warnings)

    # Formation Tops
    formation_rows = list(_iter_named_range(wb, NAMED_RANGE_FORMATION))
    if formation_rows:
        header = _header_map(formation_rows[0])
        geometry["formation_tops"] = _parse_formation_rows(formation_rows[1:], header, warnings)

    # Perforations
    perf_rows = list(_iter_named_range(wb, NAMED_RANGE_PERFORATIONS))
    if perf_rows:
        header = _header_map(perf_rows[0])
        geometry["perforations"] = _parse_perforation_rows(perf_rows[1:], header, warnings)

    # Tubing
    tubing_rows = list(_iter_named_range(wb, NAMED_RANGE_TUBING))
    if tubing_rows:
        header = _header_map(tubing_rows[0])
        geometry["tubing"] = _parse_tubing_rows(tubing_rows[1:], header, warnings)

    # Tools
    tools_rows = list(_iter_named_range(wb, NAMED_RANGE_TOOLS))
    if tools_rows:
        header = _header_map(tools_rows[0])
        geometry["tools"] = _parse_tools_rows(tools_rows[1:], header, warnings)


def _scan_geometry_rows(ws: Worksheet, geometry: dict, warnings: list[str]) -> None:
    """
    Fallback: scan Sheet 3 sequentially, detecting section headers and parsing sub-tables.
    """
    all_rows = list(ws.iter_rows())
    i = 0
    current_section: str | None = None
    current_header: dict[str, int] | None = None
    section_data_rows: list[tuple] = []

    def flush_section():
        nonlocal current_section, current_header, section_data_rows
        if current_section and current_header is not None and section_data_rows:
            _parse_section_into_geometry(
                current_section, section_data_rows, current_header, geometry, warnings
            )
        section_data_rows = []
        current_header = None

    while i < len(all_rows):
        row = all_rows[i]

        if _is_blank_row(row):
            flush_section()
            current_section = None
            i += 1
            continue

        section = _section_from_row(row)
        if section is not None and _is_section_header(row):
            # This is a section title row
            flush_section()
            current_section = section
            # Next non-blank row should be the column header row
            i += 1
            while i < len(all_rows) and _is_blank_row(all_rows[i]):
                i += 1
            if i < len(all_rows):
                current_header = _header_map(all_rows[i])
                i += 1
            continue

        # If we're inside a section, accumulate data rows
        if current_section is not None and current_header is not None:
            section_data_rows.append(row)

        i += 1

    flush_section()


def _parse_section_into_geometry(
    section: str,
    rows: list[tuple],
    header: dict[str, int],
    geometry: dict,
    warnings: list[str],
) -> None:
    parsers = {
        "casing_strings": _parse_casing_rows,
        "formation_tops": _parse_formation_rows,
        "perforations": _parse_perforation_rows,
        "tubing": _parse_tubing_rows,
        "tools": _parse_tools_rows,
    }
    parser = parsers.get(section)
    if parser:
        geometry[section].extend(parser(rows, header, warnings))


# ---------------------------------------------------------------------------
# Sub-table row parsers
# ---------------------------------------------------------------------------


def _parse_casing_rows(rows: list[tuple], header: dict[str, int], warnings: list[str]) -> list[dict]:
    """Parse Casing Record rows."""
    records = []
    patterns = {
        "string": ["string", "casing string", "name"],
        "size_in": ["size (in)", "size(in)", "size in", "od (in)", "od(in)", "od"],
        "top_ft": ["top (ft)", "top(ft)", "top ft", "top"],
        "bottom_ft": ["bottom (ft)", "bottom(ft)", "bottom ft", "bottom"],
        "hole_size_in": ["hole size (in)", "hole size(in)", "hole size", "bit size (in)", "bit size"],
        "cement_top_ft": ["cement top (ft)", "cement top(ft)", "cement top ft", "cement top"],
        "id_in": ["id (in)", "id(in)", "id in", "id"],
    }
    col = _match_columns(patterns, header)

    for row in rows:
        if _is_blank_row(row):
            continue
        vals = _row_values(row)

        def get(key, default=None):
            idx = col.get(key)
            return vals[idx] if idx is not None and idx < len(vals) else default

        string_val = _safe_str(get("string"))
        top = _safe_float(get("top_ft"))
        bottom = _safe_float(get("bottom_ft"))

        if not string_val:
            warnings.append(f"Casing Record row {row[0].row}: missing String — skipping")
            continue
        if top is None or bottom is None:
            warnings.append(f"Casing Record row {row[0].row}: missing top/bottom depth for '{string_val}' — skipping")
            continue

        records.append({
            "string": string_val,
            "size_in": _safe_float(get("size_in")),
            "top_ft": top,
            "bottom_ft": bottom,
            "hole_size_in": _safe_float(get("hole_size_in")),
            "cement_top_ft": _safe_float(get("cement_top_ft")),
            "id_in": _safe_float(get("id_in")),
        })
    return records


def _parse_formation_rows(rows: list[tuple], header: dict[str, int], warnings: list[str]) -> list[dict]:
    """Parse Formation Tops rows."""
    records = []
    patterns = {
        "formation": ["formation", "formation name", "name"],
        "depth": ["top (ft)", "top(ft)", "top ft", "depth (ft)", "depth ft", "top", "depth"],
        "base_depth": ["base (ft)", "base(ft)", "base ft", "base depth (ft)", "base"],
    }
    col = _match_columns(patterns, header)

    for row in rows:
        if _is_blank_row(row):
            continue
        vals = _row_values(row)

        def get(key, default=None):
            idx = col.get(key)
            return vals[idx] if idx is not None and idx < len(vals) else default

        formation = _safe_str(get("formation"))
        depth = _safe_float(get("depth"))

        if not formation:
            warnings.append(f"Formation Tops row {row[0].row}: missing Formation — skipping")
            continue
        if depth is None:
            warnings.append(f"Formation Tops row {row[0].row}: missing depth for '{formation}' — skipping")
            continue

        records.append({
            "formation": formation,
            "depth": depth,
            "base_depth": _safe_float(get("base_depth")),
        })
    return records


def _parse_perforation_rows(rows: list[tuple], header: dict[str, int], warnings: list[str]) -> list[dict]:
    """Parse Perforations rows."""
    records = []
    patterns = {
        "top_ft": ["top (ft)", "top(ft)", "top ft", "top"],
        "bottom_ft": ["bottom (ft)", "bottom(ft)", "bottom ft", "bottom"],
    }
    col = _match_columns(patterns, header)

    for row in rows:
        if _is_blank_row(row):
            continue
        vals = _row_values(row)

        def get(key, default=None):
            idx = col.get(key)
            return vals[idx] if idx is not None and idx < len(vals) else default

        top = _safe_float(get("top_ft"))
        bottom = _safe_float(get("bottom_ft"))

        if top is None or bottom is None:
            warnings.append(f"Perforations row {row[0].row}: missing top or bottom depth — skipping")
            continue

        records.append({"top_ft": top, "bottom_ft": bottom})
    return records


def _parse_tubing_rows(rows: list[tuple], header: dict[str, int], warnings: list[str]) -> list[dict]:
    """Parse Tubing rows."""
    records = []
    patterns = {
        "size_in": ["size (in)", "size(in)", "size in", "od (in)", "od", "size"],
        "top_ft": ["top (ft)", "top(ft)", "top ft", "top"],
        "bottom_ft": ["bottom (ft)", "bottom(ft)", "bottom ft", "bottom"],
    }
    col = _match_columns(patterns, header)

    for row in rows:
        if _is_blank_row(row):
            continue
        vals = _row_values(row)

        def get(key, default=None):
            idx = col.get(key)
            return vals[idx] if idx is not None and idx < len(vals) else default

        size = _safe_float(get("size_in"))
        top = _safe_float(get("top_ft"))
        bottom = _safe_float(get("bottom_ft"))

        if size is None or top is None or bottom is None:
            warnings.append(f"Tubing row {row[0].row}: missing required field(s) — skipping")
            continue

        records.append({"size_in": size, "top_ft": top, "bottom_ft": bottom})
    return records


def _parse_tools_rows(rows: list[tuple], header: dict[str, int], warnings: list[str]) -> list[dict]:
    """Parse Tools/Equipment rows."""
    records = []
    patterns = {
        "type": ["type", "tool type", "equipment type"],
        "top_ft": ["top (ft)", "top(ft)", "top ft", "top"],
        "bottom_ft": ["bottom (ft)", "bottom(ft)", "bottom ft", "bottom"],
        "depth_ft": ["depth (ft)", "depth(ft)", "depth ft", "depth"],
        "description": ["description", "desc", "notes"],
    }
    col = _match_columns(patterns, header)

    for row in rows:
        if _is_blank_row(row):
            continue
        vals = _row_values(row)

        def get(key, default=None):
            idx = col.get(key)
            return vals[idx] if idx is not None and idx < len(vals) else default

        tool_type = _safe_str(get("type"))
        if not tool_type:
            warnings.append(f"Tools row {row[0].row}: missing Type — skipping")
            continue

        records.append({
            "type": tool_type,
            "top_ft": _safe_float(get("top_ft")),
            "bottom_ft": _safe_float(get("bottom_ft")),
            "depth_ft": _safe_float(get("depth_ft")),
            "description": _safe_str(get("description")) or "",
        })
    return records


# ---------------------------------------------------------------------------
# Generic column matcher
# ---------------------------------------------------------------------------


def _match_columns(patterns: dict[str, list[str]], header: dict[str, int]) -> dict[str, int]:
    """
    For each logical field in patterns, find the first matching header key.
    Returns a mapping from field name → column index.
    """
    result: dict[str, int] = {}
    for field, candidates in patterns.items():
        for candidate in candidates:
            if candidate in header:
                result[field] = header[candidate]
                break
    return result


# ---------------------------------------------------------------------------
# Sheet resolution helpers
# ---------------------------------------------------------------------------


def _find_sheet(wb: openpyxl.Workbook, name: str) -> Worksheet | None:
    """Find a sheet by exact name, then by index (Sheet 2 → index 1, Sheet 3 → index 2)."""
    if name in wb.sheetnames:
        return wb[name]
    # Try index-based fallback
    index_map = {"Plug Details": 1, "Well Geometry": 2}
    idx = index_map.get(name)
    if idx is not None and len(wb.sheetnames) > idx:
        return wb.worksheets[idx]
    return None


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------


def parse_wbd_excel(file_bytes: bytes) -> dict:
    """
    Parse a WBD Excel workbook and return plug details + well geometry.

    Args:
        file_bytes: Raw bytes of an .xlsx workbook.

    Returns:
        dict with keys: "plugs", "well_geometry", "warnings"

    Raises:
        ValueError: If the file is not a valid Excel workbook or required sheets are missing.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError("File is not a valid Excel workbook") from exc

    warnings: list[str] = []

    # --- Locate required sheets ---
    ws_plugs = _find_sheet(wb, SHEET_PLUG_DETAILS)
    if ws_plugs is None:
        raise ValueError(f"Required sheet '{SHEET_PLUG_DETAILS}' not found")

    ws_geometry = _find_sheet(wb, SHEET_WELL_GEOMETRY)
    if ws_geometry is None:
        raise ValueError(f"Required sheet '{SHEET_WELL_GEOMETRY}' not found")

    # --- Parse ---
    plugs = _parse_plug_details(ws_plugs, wb, warnings)
    well_geometry = _parse_well_geometry(ws_geometry, wb, warnings)

    return {
        "plugs": plugs,
        "well_geometry": well_geometry,
        "warnings": warnings,
    }
