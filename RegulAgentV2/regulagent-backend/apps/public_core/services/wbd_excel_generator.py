"""
WBD Excel Generator
====================
Generates a 4-sheet Excel workbook representing a Wellbore Diagram (WBD)
from reconciliation and well geometry data.

Entry point:
    generate_wbd_excel(data: dict) -> BytesIO

Sheets:
    1. AS PLUGGED WBD    — Visual cell-art schematic with casing walls, plugs, formations
    2. Plug Details      — Tabular plug data with data validation
    3. Well Geometry     — Sub-tables for casing, formations, perfs, tubing, tools
    4. Well Bore Icons   — Color reference key
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="DBEAFE")
BOLD_FONT = Font(bold=True)

PLUG_COLORS = {
    "cement": "22C55E",
    "bridge_plug": "FCD34D",
    "cast_iron_bridge_plug": "FCD34D",
}
PLUG_DEFAULT_COLOR = "22C55E"

PERF_COLOR = "F97316"

# New schematic color constants
CASING_WALL_COLOR = "6B7280"       # Dark gray for casing steel walls
CASING_CEMENT_ANNULAR = "D1D5DB"  # Light gray for cement behind casing


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _apply_header_style(cell, text: str) -> None:
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


def _apply_cell_border(cell) -> None:
    cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Sheet 1: AS PLUGGED WBD — Visual cell-art schematic
# ---------------------------------------------------------------------------

# Diagram layout constants
DIAGRAM_START_ROW = 9
NUM_DIAGRAM_ROWS = 50

# Wall column pairs (outermost to innermost)
WALL_PAIRS = [
    ("C", "M"),  # Outermost (surface)
    ("D", "L"),  # Second (intermediate)
    ("E", "K"),  # Third (production)
    ("F", "J"),  # Fourth (liner)
]

# Column widths for the schematic sheet
SCHEMATIC_COLUMN_WIDTHS = {
    "A": 6,
    "B": 2.5,
    "C": 4,
    "D": 4,
    "E": 4,
    "F": 3,
    "G": 1.5,
    "H": 8,
    "I": 1.5,
    "J": 3,
    "K": 4,
    "L": 4,
    "M": 4,
    "N": 12,
    "O": 20,
    "Y": 16,
    "Z": 8,
}


def _col_index(col_letter: str) -> int:
    """Convert a column letter (A-Z) to 1-based column index."""
    return ord(col_letter.upper()) - ord("A") + 1


def _depth_to_row(
    depth_ft: float,
    max_depth: float,
    start_row: int = DIAGRAM_START_ROW,
    num_rows: int = NUM_DIAGRAM_ROWS,
) -> int:
    """Map a depth in feet to a worksheet row number, proportionally."""
    if max_depth <= 0:
        return start_row
    frac = depth_ft / max_depth
    return start_row + round(frac * num_rows)


def _col_range(left_col: str, right_col: str) -> list[str]:
    """Return all column letters from left_col to right_col inclusive."""
    left_idx = _col_index(left_col)
    right_idx = _col_index(right_col)
    return [get_column_letter(i) for i in range(left_idx, right_idx + 1)]


def _build_wellbore_diagram(ws, data: dict) -> None:
    ws.title = "AS PLUGGED WBD"

    # ---- Column widths ----
    for col_letter, width in SCHEMATIC_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ---- Row heights for header rows (1-8) ----
    for r in range(1, 9):
        ws.row_dimensions[r].height = 18

    header = data.get("well_header", {})
    geometry = data.get("well_geometry", {})
    comparisons = data.get("comparisons", [])

    # =========================================================================
    # HEADER SECTION (Rows 1-8)
    # =========================================================================

    # Row 1: blank
    # Row 2
    ws.cell(row=2, column=_col_index("A"), value="Author:").font = BOLD_FONT
    ws.cell(row=2, column=_col_index("B"), value=header.get("operator", ""))
    ws.cell(row=2, column=_col_index("O"), value="Well No:").font = BOLD_FONT
    ws.cell(row=2, column=_col_index("P"), value=header.get("name", ""))

    # Row 3
    ws.cell(row=3, column=_col_index("A"), value="Well Name:").font = BOLD_FONT
    ws.cell(row=3, column=_col_index("B"), value=header.get("name", ""))
    ws.cell(row=3, column=_col_index("O"), value="API #:").font = BOLD_FONT
    ws.cell(row=3, column=_col_index("P"), value=header.get("api_number", ""))

    # Row 4
    ws.cell(row=4, column=_col_index("A"), value="Field/Pool:").font = BOLD_FONT
    ws.cell(row=4, column=_col_index("B"), value=header.get("field", ""))
    ws.cell(row=4, column=_col_index("O"), value="Location:").font = BOLD_FONT
    ws.cell(row=4, column=_col_index("P"), value="")

    # Row 5
    ws.cell(row=5, column=_col_index("A"), value="County:").font = BOLD_FONT
    ws.cell(row=5, column=_col_index("B"), value=header.get("county", ""))

    # Row 6
    ws.cell(row=6, column=_col_index("A"), value="State:").font = BOLD_FONT
    ws.cell(row=6, column=_col_index("B"), value=data.get("jurisdiction", ""))
    ws.cell(row=6, column=_col_index("O"), value="GL:").font = BOLD_FONT
    ws.cell(row=6, column=_col_index("P"), value="")

    # Row 7
    ws.cell(row=7, column=_col_index("A"), value="Spud Date:").font = BOLD_FONT
    ws.cell(row=7, column=_col_index("B"), value=header.get("spud_date", ""))

    # Row 8: blank separator

    # ---- Casing Record Table (right side, cols S-Z, rows 1-2+) ----
    casing_record_headers = ["Description", "OD", "Grade", "Weight", "Depth", "Hole", "Cmt Sx", "TOC"]
    casing_start_col = _col_index("S")
    for col_offset, hdr_text in enumerate(casing_record_headers):
        cell = ws.cell(row=1, column=casing_start_col + col_offset, value=hdr_text)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_offset, cs in enumerate(geometry.get("casing_strings", []), start=1):
        row_num = 1 + row_offset
        cs_values = [
            cs.get("string", ""),
            cs.get("size_in", ""),
            cs.get("grade", ""),
            cs.get("weight_lb_ft", ""),
            cs.get("bottom_ft", ""),
            cs.get("hole_size_in", ""),
            cs.get("cement_sacks", ""),
            cs.get("cement_top_ft", ""),
        ]
        for col_offset, val in enumerate(cs_values):
            cell = ws.cell(row=row_num, column=casing_start_col + col_offset, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # =========================================================================
    # COMPUTE max_depth for proportional mapping
    # =========================================================================

    depth_candidates: list[float] = [0.0]

    for cs in geometry.get("casing_strings", []):
        if cs.get("bottom_ft") is not None:
            depth_candidates.append(float(cs["bottom_ft"]))

    for ft in geometry.get("formation_tops", []):
        if ft.get("base_depth") is not None:
            depth_candidates.append(float(ft["base_depth"]))
        if ft.get("depth") is not None:
            depth_candidates.append(float(ft["depth"]))

    for comp in comparisons:
        if comp.get("actual_bottom_ft") is not None:
            depth_candidates.append(float(comp["actual_bottom_ft"]))
        if comp.get("actual_top_ft") is not None:
            depth_candidates.append(float(comp["actual_top_ft"]))

    for p in geometry.get("perforations", []):
        if p.get("bottom_ft") is not None:
            depth_candidates.append(float(p["bottom_ft"]))

    max_depth = max(depth_candidates) if depth_candidates else 1000.0

    # =========================================================================
    # PRE-PROCESS geometry
    # =========================================================================

    # Sort casing strings by hole_size_in descending (outermost first)
    raw_casings = geometry.get("casing_strings", [])
    sorted_casings = sorted(
        raw_casings,
        key=lambda cs: float(cs.get("hole_size_in") or 0),
        reverse=True,
    )

    # Assign wall pairs; casings beyond 4 entries don't get walls
    casing_wall_assignments: list[dict] = []
    for idx, cs in enumerate(sorted_casings):
        if idx < len(WALL_PAIRS):
            left_col, right_col = WALL_PAIRS[idx]
        else:
            left_col, right_col = None, None
        casing_wall_assignments.append({
            "string": cs.get("string", ""),
            "top_ft": float(cs.get("top_ft", 0)),
            "bottom_ft": float(cs.get("bottom_ft", 0)),
            "cement_top_ft": float(cs["cement_top_ft"]) if cs.get("cement_top_ft") is not None else None,
            "hole_size_in": float(cs.get("hole_size_in") or 0),
            "left_col": left_col,
            "right_col": right_col,
        })

    # Determine interior columns (between innermost wall pair)
    if casing_wall_assignments:
        assigned = [c for c in casing_wall_assignments if c["left_col"] is not None]
        if assigned:
            innermost = assigned[-1]
            inner_left = innermost["left_col"]
            inner_right = innermost["right_col"]
            # Interior = columns strictly between inner_left and inner_right
            interior_start = _col_index(inner_left) + 1
            interior_end = _col_index(inner_right) - 1
            interior_cols = [get_column_letter(i) for i in range(interior_start, interior_end + 1)]
        else:
            interior_cols = ["G", "H", "I"]
    else:
        interior_cols = ["G", "H", "I"]

    # Plugs from comparisons
    plugs: list[dict] = []
    for comp in comparisons:
        ptype = comp.get("actual_type") or "cement"
        top = comp.get("actual_top_ft")
        bottom = comp.get("actual_bottom_ft")
        if top is not None and bottom is not None:
            plugs.append({
                "plug_number": comp.get("plug_number", 0),
                "type": ptype,
                "top_ft": float(top),
                "bottom_ft": float(bottom),
                "sacks": comp.get("actual_sacks"),
                "cement_class": comp.get("actual_cement_class"),
            })

    # Formation tops
    formations: list[dict] = []
    for ft in geometry.get("formation_tops", []):
        if ft.get("depth") is not None:
            formations.append({
                "name": ft.get("formation", "Unknown"),
                "top_ft": float(ft["depth"]),
            })

    # Perforations
    perfs: list[dict] = []
    for p in geometry.get("perforations", []):
        if p.get("top_ft") is not None and p.get("bottom_ft") is not None:
            perfs.append({
                "top_ft": float(p["top_ft"]),
                "bottom_ft": float(p["bottom_ft"]),
            })

    # =========================================================================
    # SET DIAGRAM ROW HEIGHTS
    # =========================================================================
    for r in range(DIAGRAM_START_ROW, DIAGRAM_START_ROW + NUM_DIAGRAM_ROWS + 1):
        ws.row_dimensions[r].height = 15

    # =========================================================================
    # ELEMENT 1: DEPTH LABELS (col A)
    # =========================================================================
    # Collect all key depths; deduplicate to avoid overlapping labels
    key_depths: list[float] = [0.0, max_depth]

    for cwa in casing_wall_assignments:
        key_depths.append(cwa["bottom_ft"])
        if cwa["cement_top_ft"] is not None:
            key_depths.append(cwa["cement_top_ft"])

    for plug in plugs:
        key_depths.append(plug["top_ft"])
        key_depths.append(plug["bottom_ft"])

    for formation in formations:
        key_depths.append(formation["top_ft"])

    for perf in perfs:
        key_depths.append(perf["top_ft"])
        key_depths.append(perf["bottom_ft"])

    # Map each key depth to a row, then write (first-write-wins per row)
    used_depth_label_rows: set[int] = set()
    for depth_val in sorted(set(key_depths)):
        row_num = _depth_to_row(depth_val, max_depth)
        if row_num not in used_depth_label_rows:
            cell = ws.cell(row=row_num, column=_col_index("A"), value=int(depth_val))
            cell.font = Font(size=8)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            used_depth_label_rows.add(row_num)

    # =========================================================================
    # ELEMENT 2: CASING WALLS (fill left_col and right_col cells row-by-row)
    # =========================================================================
    for cwa in casing_wall_assignments:
        if cwa["left_col"] is None:
            continue
        top_row = _depth_to_row(cwa["top_ft"], max_depth)
        bottom_row = _depth_to_row(cwa["bottom_ft"], max_depth)
        wall_fill = _make_fill(CASING_WALL_COLOR)
        for r in range(top_row, bottom_row + 1):
            ws.cell(row=r, column=_col_index(cwa["left_col"])).fill = wall_fill
            ws.cell(row=r, column=_col_index(cwa["right_col"])).fill = wall_fill

    # =========================================================================
    # ELEMENT 3: CEMENT BEHIND CASING (annular columns between casing pairs)
    # =========================================================================
    assigned_casings = [c for c in casing_wall_assignments if c["left_col"] is not None]
    for idx, cwa in enumerate(assigned_casings):
        if cwa["cement_top_ft"] is None:
            continue
        # Skip outermost — no annular space outside it
        if idx == 0:
            continue
        outer_cwa = assigned_casings[idx - 1]

        cement_top_row = _depth_to_row(cwa["cement_top_ft"], max_depth)
        cement_bottom_row = _depth_to_row(cwa["bottom_ft"], max_depth)

        # Annular columns: between outer casing's left_col and this casing's left_col
        # and between this casing's right_col and outer casing's right_col
        outer_left_idx = _col_index(outer_cwa["left_col"])
        inner_left_idx = _col_index(cwa["left_col"])
        outer_right_idx = _col_index(outer_cwa["right_col"])
        inner_right_idx = _col_index(cwa["right_col"])

        annular_cols = (
            list(range(outer_left_idx + 1, inner_left_idx))
            + list(range(inner_right_idx + 1, outer_right_idx))
        )
        cement_fill = _make_fill(CASING_CEMENT_ANNULAR)
        for r in range(cement_top_row, cement_bottom_row + 1):
            for col_idx in annular_cols:
                ws.cell(row=r, column=col_idx).fill = cement_fill

    # =========================================================================
    # ELEMENT 4: PLUGS (merge interior columns across depth rows)
    # =========================================================================
    if interior_cols:
        first_interior_col = interior_cols[0]
        last_interior_col = interior_cols[-1]
        first_interior_idx = _col_index(first_interior_col)
        last_interior_idx = _col_index(last_interior_col)
    else:
        first_interior_idx = _col_index("G")
        last_interior_idx = _col_index("I")
        first_interior_col = "G"
        last_interior_col = "I"

    for plug in plugs:
        top_row = _depth_to_row(plug["top_ft"], max_depth)
        bottom_row = _depth_to_row(plug["bottom_ft"], max_depth)

        # Clamp rows to diagram range
        top_row = max(top_row, DIAGRAM_START_ROW)
        bottom_row = min(bottom_row, DIAGRAM_START_ROW + NUM_DIAGRAM_ROWS)

        if top_row > bottom_row:
            bottom_row = top_row

        plug_color = PLUG_COLORS.get(plug["type"], PLUG_DEFAULT_COLOR)
        plug_fill = _make_fill(plug_color)

        # Build merge range string: first_interior_col + top_row : last_interior_col + bottom_row
        merge_ref = (
            f"{first_interior_col}{top_row}:{last_interior_col}{bottom_row}"
        )
        try:
            ws.merge_cells(merge_ref)
        except Exception:
            pass  # Overlapping merges are non-fatal

        anchor_cell = ws.cell(row=top_row, column=first_interior_idx)
        anchor_cell.fill = plug_fill
        anchor_cell.value = f"Plug #{plug['plug_number']}"
        anchor_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        anchor_cell.font = Font(bold=True, size=9)

        # Fill all cells in the merged range with the plug color
        for r in range(top_row, bottom_row + 1):
            for c in range(first_interior_idx, last_interior_idx + 1):
                ws.cell(row=r, column=c).fill = plug_fill

    # =========================================================================
    # ELEMENT 5: TEXT ANNOTATIONS (cols N-O) for plugs
    # =========================================================================
    for plug in plugs:
        top_row = _depth_to_row(plug["top_ft"], max_depth)
        bottom_row = _depth_to_row(plug["bottom_ft"], max_depth)
        mid_row = (top_row + bottom_row) // 2
        mid_row = max(mid_row, DIAGRAM_START_ROW)

        desc = f"Plug #{plug['plug_number']}: {plug['type']} from {int(plug['top_ft'])}' to {int(plug['bottom_ft'])}'"
        if plug.get("sacks") and plug.get("cement_class"):
            desc += f" ({plug['sacks']} sxs class {plug['cement_class']})"
        elif plug.get("sacks"):
            desc += f" ({plug['sacks']} sxs)"

        ws.cell(row=mid_row, column=_col_index("N"), value=plug["plug_number"])
        ann_cell = ws.cell(row=mid_row, column=_col_index("O"), value=desc)
        ann_cell.font = Font(size=9)
        ann_cell.alignment = Alignment(wrap_text=True, vertical="center")

    # =========================================================================
    # ELEMENT 6: FORMATION TOPS (cols Y-Z)
    # =========================================================================
    for formation in formations:
        row_num = _depth_to_row(formation["top_ft"], max_depth)
        row_num = max(row_num, DIAGRAM_START_ROW)

        name_cell = ws.cell(row=row_num, column=_col_index("Y"), value=formation["name"])
        name_cell.font = Font(size=9)
        name_cell.alignment = Alignment(vertical="center")

        depth_cell = ws.cell(row=row_num, column=_col_index("Z"), value=int(formation["top_ft"]))
        depth_cell.font = Font(size=9)
        depth_cell.alignment = Alignment(horizontal="right", vertical="center")

    # =========================================================================
    # ELEMENT 7: PERFORATIONS (fill interior columns)
    # =========================================================================
    perf_fill = _make_fill(PERF_COLOR)
    for perf in perfs:
        top_row = _depth_to_row(perf["top_ft"], max_depth)
        bottom_row = _depth_to_row(perf["bottom_ft"], max_depth)
        top_row = max(top_row, DIAGRAM_START_ROW)
        bottom_row = min(bottom_row, DIAGRAM_START_ROW + NUM_DIAGRAM_ROWS)
        for r in range(top_row, bottom_row + 1):
            for c in range(first_interior_idx, last_interior_idx + 1):
                ws.cell(row=r, column=c).fill = perf_fill

    # =========================================================================
    # ELEMENT 8: BOTTOM LABEL (TD)
    # =========================================================================
    td_row = _depth_to_row(max_depth, max_depth)
    td_row = min(td_row, DIAGRAM_START_ROW + NUM_DIAGRAM_ROWS)
    ws.cell(row=td_row, column=_col_index("H"), value="TD").font = Font(bold=True, size=9)


# ---------------------------------------------------------------------------
# Sheet 2: Plug Details
# ---------------------------------------------------------------------------

def _build_plug_details(ws, data: dict) -> None:
    ws.title = "Plug Details"

    # Row 1: title header
    ws.merge_cells("A1:J1")
    title_cell = ws.cell(row=1, column=1, value="Plug Details")
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = THIN_BORDER

    # Row 2: column headers
    col_headers = [
        "Plug #", "Type", "Top (ft)", "Bottom (ft)", "Sacks",
        "Cement Class", "Tagged Depth (ft)", "Placement Method",
        "WOC Hours", "WOC Tagged",
    ]
    for col_idx, text in enumerate(col_headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        _apply_header_style(cell, text)

    comparisons = data.get("comparisons", [])
    for row_idx, comp in enumerate(comparisons, start=3):
        values = [
            comp.get("plug_number"),
            comp.get("actual_type"),
            comp.get("actual_top_ft"),
            comp.get("actual_bottom_ft"),
            comp.get("actual_sacks"),
            comp.get("actual_cement_class"),
            comp.get("actual_tagged_depth_ft"),
            comp.get("actual_placement_method"),
            comp.get("actual_woc_hours"),
            comp.get("actual_woc_tagged"),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    last_row = 2 + len(comparisons)

    # Data validation: Type (col B)
    dv_type = DataValidation(
        type="list",
        formula1='"cement,bridge_plug,cast_iron_bridge_plug"',
        allow_blank=True,
        showErrorMessage=True,
    )
    ws.add_data_validation(dv_type)
    if comparisons:
        dv_type.add(f"B3:B{last_row}")

    # Data validation: Placement Method (col H)
    dv_method = DataValidation(
        type="list",
        formula1='"pump_and_plug,balanced,dump_bailer"',
        allow_blank=True,
        showErrorMessage=True,
    )
    ws.add_data_validation(dv_method)
    if comparisons:
        dv_method.add(f"H3:H{last_row}")

    # Named range
    from openpyxl.workbook.defined_name import DefinedName
    ref = f"'Plug Details'!$A$2:$J${last_row}"
    ws.parent.defined_names["PlugDetails"] = DefinedName("PlugDetails", attr_text=ref)

    # Column widths
    widths = [8, 22, 10, 12, 8, 14, 16, 20, 12, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 3: Well Geometry
# ---------------------------------------------------------------------------

def _write_sub_table(
    ws,
    start_row: int,
    title: str,
    col_headers: list[str],
    rows: list[list],
    named_range_name: str,
) -> int:
    """Write a sub-table and return the next available row after it."""
    num_cols = len(col_headers)
    end_col_letter = get_column_letter(num_cols)

    # Title row
    ws.merge_cells(f"A{start_row}:{end_col_letter}{start_row}")
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = HEADER_FONT
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center")
    title_cell.border = THIN_BORDER

    # Header row
    header_row = start_row + 1
    for col_idx, text in enumerate(col_headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        _apply_header_style(cell, text)

    # Data rows
    data_start = header_row + 1
    for row_offset, row_data in enumerate(rows):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=data_start + row_offset, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    last_data_row = data_start + len(rows) - 1 if rows else header_row

    # Named range
    from openpyxl.workbook.defined_name import DefinedName
    ref = f"'Well Geometry'!$A${header_row}:${end_col_letter}${last_data_row}"
    try:
        ws.parent.defined_names[named_range_name] = DefinedName(named_range_name, attr_text=ref)
    except Exception:
        pass  # Named range conflicts are non-fatal

    return last_data_row + 3  # +2 blank rows, then next table starts 1 row after


def _build_well_geometry(ws, data: dict) -> None:
    ws.title = "Well Geometry"

    geometry = data.get("well_geometry", {})
    current_row = 1

    # --- Casing Record ---
    casing_rows = []
    for cs in geometry.get("casing_strings", []):
        casing_rows.append([
            cs.get("string"),
            cs.get("size_in"),
            cs.get("top_ft"),
            cs.get("bottom_ft"),
            cs.get("hole_size_in"),
            cs.get("cement_top_ft"),
            cs.get("id_in"),
        ])
    current_row = _write_sub_table(
        ws, current_row, "Casing Record",
        ["String", "Size (in)", "Top (ft)", "Bottom (ft)", "Hole Size (in)", "Cement Top (ft)", "ID (in)"],
        casing_rows, "CasingRecord",
    )

    # --- Formation Tops ---
    formation_rows = []
    for ft in geometry.get("formation_tops", []):
        formation_rows.append([
            ft.get("formation"),
            ft.get("depth"),
            ft.get("base_depth"),
        ])
    current_row = _write_sub_table(
        ws, current_row, "Formation Tops",
        ["Formation", "Top (ft)", "Base (ft)"],
        formation_rows, "FormationTops",
    )

    # --- Perforations ---
    perf_rows = [
        [p.get("top_ft"), p.get("bottom_ft")]
        for p in geometry.get("perforations", [])
    ]
    current_row = _write_sub_table(
        ws, current_row, "Perforations",
        ["Top (ft)", "Bottom (ft)"],
        perf_rows, "Perforations",
    )

    # --- Tubing ---
    tubing_rows = [
        [t.get("size_in"), t.get("top_ft"), t.get("bottom_ft")]
        for t in geometry.get("tubing", [])
    ]
    current_row = _write_sub_table(
        ws, current_row, "Tubing",
        ["Size (in)", "Top (ft)", "Bottom (ft)"],
        tubing_rows, "Tubing",
    )

    # --- Tools/Equipment ---
    tool_rows = [
        [t.get("type"), t.get("top_ft"), t.get("bottom_ft"), t.get("depth_ft"), t.get("description")]
        for t in geometry.get("tools", [])
    ]
    _write_sub_table(
        ws, current_row, "Tools/Equipment",
        ["Type", "Top (ft)", "Bottom (ft)", "Depth (ft)", "Description"],
        tool_rows, "ToolsEquipment",
    )

    # Column widths
    col_widths = [22, 12, 10, 12, 12, 14, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet 4: Legend
# ---------------------------------------------------------------------------

def _build_legend(ws, data: dict) -> None:
    ws.title = "Well Bore Icons"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30

    # Row 1: Title
    ws.merge_cells("A1:B1")
    title = ws.cell(row=1, column=1, value="Well Bore Icons")
    title.font = Font(bold=True, size=14)
    title.alignment = Alignment(horizontal="center")

    # Core entries only — no formation-specific colors
    entries = [
        ("22C55E", "Cement Plug"),
        ("FCD34D", "Bridge Plug"),
        (CASING_WALL_COLOR, "Casing Steel"),
        (CASING_CEMENT_ANNULAR, "Cement Behind Casing"),
        (PERF_COLOR, "Perforations"),
    ]

    for row_idx, (color, description) in enumerate(entries, start=3):
        swatch = ws.cell(row=row_idx, column=1)
        swatch.fill = _make_fill(color)
        swatch.border = THIN_BORDER

        label = ws.cell(row=row_idx, column=2, value=description)
        label.alignment = Alignment(vertical="center")
        label.border = THIN_BORDER

        ws.row_dimensions[row_idx].height = 18


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def generate_wbd_excel(data: dict) -> BytesIO:
    """
    Build a 4-sheet Excel workbook from WBD and reconciliation data.

    Args:
        data: Dictionary containing well_header, jurisdiction, comparisons,
              and well_geometry keys (see module docstring for full schema).

    Returns:
        BytesIO buffer positioned at offset 0, ready for download or storage.
    """
    wb = Workbook()

    # openpyxl creates a default sheet; repurpose it as the first sheet
    ws1 = wb.active
    _build_wellbore_diagram(ws1, data)

    ws2 = wb.create_sheet()
    _build_plug_details(ws2, data)

    ws3 = wb.create_sheet()
    _build_well_geometry(ws3, data)

    ws4 = wb.create_sheet()
    _build_legend(ws4, data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
