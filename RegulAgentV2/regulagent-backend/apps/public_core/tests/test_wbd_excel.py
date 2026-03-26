"""
Tests for WBD Excel Generator and Parser — round-trip verification.

No database access required; these are pure unit tests on the service functions.
"""
import pytest
import openpyxl
from io import BytesIO

from apps.public_core.services.wbd_excel_generator import generate_wbd_excel
from apps.public_core.services.wbd_excel_parser import parse_wbd_excel


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

SAMPLE_DATA = {
    "well_header": {
        "name": "Test Well #1",
        "api_number": "42123456789012",
        "operator": "Test Operator",
        "field": "Test Field",
    },
    "jurisdiction": "TX",
    "comparisons": [
        {
            "plug_number": 1,
            "actual_type": "cement",
            "actual_top_ft": 5000.0,
            "actual_bottom_ft": 5100.0,
            "actual_sacks": 50.0,
            "actual_cement_class": "H",
            "actual_tagged_depth_ft": 5005.0,
            "actual_placement_method": "pump_and_plug",
            "actual_woc_hours": 8.0,
            "actual_woc_tagged": True,
        },
        {
            "plug_number": 2,
            "actual_type": "bridge_plug",
            "actual_top_ft": 3000.0,
            "actual_bottom_ft": 3010.0,
            "actual_sacks": None,
            "actual_cement_class": None,
            "actual_tagged_depth_ft": 3000.0,
            "actual_placement_method": None,
            "actual_woc_hours": None,
            "actual_woc_tagged": None,
        },
    ],
    "well_geometry": {
        "casing_strings": [
            {"string": "Surface", "size_in": 8.625, "top_ft": 0.0, "bottom_ft": 500.0, "hole_size_in": 12.25, "cement_top_ft": None, "id_in": None},
            {"string": "Production", "size_in": 4.5, "top_ft": 0.0, "bottom_ft": 6000.0, "hole_size_in": 7.875, "cement_top_ft": 4500.0, "id_in": None},
        ],
        "formation_tops": [
            {"formation": "Austin Chalk", "depth": 2000.0, "base_depth": 3500.0},
            {"formation": "Eagle Ford", "depth": 3500.0, "base_depth": 5000.0},
        ],
        "perforations": [
            {"top_ft": 5500.0, "bottom_ft": 5600.0},
        ],
        "tubing": [],
        "tools": [],
    },
}

EMPTY_GEOMETRY_DATA = {
    "well_header": {"name": "Empty Well", "api_number": "42000000000000", "operator": "Test", "field": "None"},
    "jurisdiction": "NM",
    "comparisons": [],
    "well_geometry": {
        "casing_strings": [],
        "formation_tops": [],
        "perforations": [],
        "tubing": [],
        "tools": [],
    },
}

DEEP_WELL_DATA = {
    "well_header": {"name": "Deep Well", "api_number": "42999999999999", "operator": "Deep Op", "field": "Deep"},
    "jurisdiction": "TX",
    "comparisons": [
        {
            "plug_number": 1,
            "actual_type": "cement",
            "actual_top_ft": 14000.0,
            "actual_bottom_ft": 14100.0,
            "actual_sacks": 100.0,
            "actual_cement_class": "H",
            "actual_tagged_depth_ft": 14005.0,
            "actual_placement_method": "balanced",
            "actual_woc_hours": 12.0,
            "actual_woc_tagged": True,
        },
    ],
    "well_geometry": {
        "casing_strings": [
            {"string": "Surface", "size_in": 13.375, "top_ft": 0.0, "bottom_ft": 2000.0, "hole_size_in": 17.5},
            {"string": "Intermediate", "size_in": 9.625, "top_ft": 0.0, "bottom_ft": 8000.0, "hole_size_in": 12.25},
            {"string": "Production", "size_in": 5.5, "top_ft": 0.0, "bottom_ft": 15000.0, "hole_size_in": 8.5},
        ],
        "formation_tops": [
            {"formation": "Wolfcamp", "depth": 5000.0, "base_depth": 10000.0},
            {"formation": "Bone Spring", "depth": 10000.0, "base_depth": 14000.0},
        ],
        "perforations": [{"top_ft": 14500.0, "bottom_ft": 14800.0}],
        "tubing": [{"size_in": 2.875, "top_ft": 0.0, "bottom_ft": 14000.0}],
        "tools": [{"type": "Packer", "top_ft": 13900.0, "bottom_ft": 13950.0, "depth_ft": 13900.0, "description": "Retrievable packer"}],
    },
}


# ---------------------------------------------------------------------------
# Generator Tests
# ---------------------------------------------------------------------------

class TestWbdExcelGenerator:
    """Tests for generate_wbd_excel()."""

    def test_generate_basic_workbook(self):
        """Verify a workbook with 4 sheets is produced."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        assert isinstance(buf, BytesIO)
        wb = openpyxl.load_workbook(buf)
        assert len(wb.sheetnames) == 4
        assert wb.sheetnames[0] == "AS PLUGGED WBD"
        assert wb.sheetnames[1] == "Plug Details"
        assert wb.sheetnames[2] == "Well Geometry"
        assert wb.sheetnames[3] == "Well Bore Icons"

    def test_generate_plug_details_sheet(self):
        """Verify Sheet 2 has correct columns and data rows."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        wb = openpyxl.load_workbook(buf)
        ws = wb["Plug Details"]

        # Check header row (row 2)
        headers = [ws.cell(row=2, column=c).value for c in range(1, 11)]
        assert "Plug #" in headers
        assert "Type" in headers
        assert "Top (ft)" in headers

        # Check data: 2 plugs → rows 3 and 4
        assert ws.cell(row=3, column=1).value == 1  # Plug #1
        assert ws.cell(row=3, column=2).value == "cement"
        assert ws.cell(row=3, column=3).value == 5000.0
        assert ws.cell(row=4, column=1).value == 2  # Plug #2
        assert ws.cell(row=4, column=2).value == "bridge_plug"

    def test_generate_geometry_sheet_has_sub_tables(self):
        """Verify workbook has named ranges for geometry sub-tables."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        wb = openpyxl.load_workbook(buf)

        # Check named ranges exist (workbook-level)
        assert wb.defined_names.get("CasingRecord") is not None
        assert wb.defined_names.get("FormationTops") is not None
        assert wb.defined_names.get("PlugDetails") is not None

    def test_generate_visual_grid_has_depths(self):
        """Verify Sheet 1 header section and diagram start row depth label."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        wb = openpyxl.load_workbook(buf)
        ws = wb["AS PLUGGED WBD"]

        # Header section: row 2 col A should have "Author:" label
        assert ws.cell(row=2, column=1).value == "Author:"

        # Row 3 col A should have "Well Name:" label
        assert ws.cell(row=3, column=1).value == "Well Name:"

        # Row 3 col B should have the actual well name value
        assert ws.cell(row=3, column=2).value == SAMPLE_DATA["well_header"]["name"]

        # DIAGRAM_START_ROW = 9; depth label at row 9, col A should be 0
        assert ws.cell(row=9, column=1).value == 0

    def test_generate_legend_sheet(self):
        """Verify Sheet 4 (Well Bore Icons) has legend entries."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        wb = openpyxl.load_workbook(buf)
        ws = wb["Well Bore Icons"]

        assert ws.cell(row=1, column=1).value == "Well Bore Icons"
        # Should have entries for cement plug, casing steel, and perforations
        legend_texts = [ws.cell(row=r, column=2).value for r in range(3, 15) if ws.cell(row=r, column=2).value]
        assert any("Cement Plug" in t for t in legend_texts)
        assert any("Casing Steel" in t for t in legend_texts)
        assert any("Perforations" in t for t in legend_texts)

    def test_proportional_depth_mapping(self):
        """Verify _depth_to_row produces proportional row positions."""
        from apps.public_core.services.wbd_excel_generator import _depth_to_row
        # For max_depth=10000, midpoint should map to midpoint row
        mid_row = _depth_to_row(5000, 10000, start_row=9, num_rows=50)
        assert mid_row == 34  # 9 + round(0.5 * 50)
        # Zero depth → start row
        assert _depth_to_row(0, 10000) == 9
        # Max depth → start + num_rows
        assert _depth_to_row(10000, 10000) == 59
        # Edge: max_depth=0
        assert _depth_to_row(500, 0) == 9

    def test_dynamic_casing_wall_assignment(self):
        """Verify casings are assigned wall columns by hole size (largest first)."""
        buf = generate_wbd_excel(DEEP_WELL_DATA)
        wb = openpyxl.load_workbook(buf)
        ws = wb["AS PLUGGED WBD"]
        # DEEP_WELL_DATA has 3 casings: Surface(17.5"), Intermediate(12.25"), Production(8.5")
        # Surface → C/M walls, Intermediate → D/L walls, Production → E/K walls
        # At depth 0 (row 9), surface casing wall at col C should have a fill
        cell_c = ws.cell(row=9, column=3)  # C9
        assert cell_c.fill.fgColor.rgb is not None  # Has fill (casing wall)

    def test_casing_record_in_header(self):
        """Verify casing record table appears in cols S-Z."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        wb = openpyxl.load_workbook(buf)
        ws = wb["AS PLUGGED WBD"]
        # Row 1, col S should have "Description" header
        assert ws.cell(row=1, column=19).value == "Description"
        # Row 2 should have first casing string name
        assert ws.cell(row=2, column=19).value is not None

    def test_plug_annotations(self):
        """Verify plug annotations appear in cols N-O."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        wb = openpyxl.load_workbook(buf)
        ws = wb["AS PLUGGED WBD"]
        # Find annotation text in col O (column 15) in diagram rows
        found_annotation = False
        for r in range(9, 60):
            val = ws.cell(row=r, column=15).value
            if val and "Plug #" in str(val):
                found_annotation = True
                break
        assert found_annotation, "No plug annotation found in col O"

    def test_formation_tops_in_cols_yz(self):
        """Verify formation tops appear in cols Y-Z."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        wb = openpyxl.load_workbook(buf)
        ws = wb["AS PLUGGED WBD"]
        # Find formation name in col Y (column 25)
        found_formation = False
        for r in range(9, 60):
            val = ws.cell(row=r, column=25).value
            if val and "Austin" in str(val):
                found_formation = True
                break
        assert found_formation, "No formation top found in col Y"

    def test_single_casing_well(self):
        """Well with only 1 casing should still render."""
        data = {
            **SAMPLE_DATA,
            "well_geometry": {
                **SAMPLE_DATA["well_geometry"],
                "casing_strings": [
                    {"string": "Surface", "size_in": 8.625, "top_ft": 0.0, "bottom_ft": 500.0, "hole_size_in": 12.25, "cement_top_ft": None},
                ],
            },
        }
        buf = generate_wbd_excel(data)
        wb = openpyxl.load_workbook(buf)
        assert wb.sheetnames[0] == "AS PLUGGED WBD"

    def test_no_formations(self):
        """Well with no formation tops should still render."""
        data = {
            **SAMPLE_DATA,
            "well_geometry": {
                **SAMPLE_DATA["well_geometry"],
                "formation_tops": [],
            },
        }
        buf = generate_wbd_excel(data)
        wb = openpyxl.load_workbook(buf)
        assert wb.sheetnames[0] == "AS PLUGGED WBD"

    def test_generate_empty_data(self):
        """Generate with empty comparisons/geometry — should not error."""
        buf = generate_wbd_excel(EMPTY_GEOMETRY_DATA)
        wb = openpyxl.load_workbook(buf)
        assert len(wb.sheetnames) == 4

    def test_generate_deep_well_tick_intervals(self):
        """Deep wells use proportional depth mapping; verify start label and TD marker."""
        buf = generate_wbd_excel(DEEP_WELL_DATA)
        wb = openpyxl.load_workbook(buf)
        ws = wb["AS PLUGGED WBD"]

        # DIAGRAM_START_ROW = 9; depth label at row 9, col A should be 0
        assert ws.cell(row=9, column=1).value == 0

        # There should be a "TD" label somewhere in col H (column 8) within diagram rows
        found_td = False
        for r in range(9, 60):
            val = ws.cell(row=r, column=8).value
            if val and "TD" in str(val):
                found_td = True
                break
        assert found_td, "No TD label found in col H of diagram rows"


# ---------------------------------------------------------------------------
# Parser Tests
# ---------------------------------------------------------------------------

class TestWbdExcelParser:
    """Tests for parse_wbd_excel()."""

    def _generate_bytes(self, data=None):
        """Helper: generate Excel bytes from data."""
        buf = generate_wbd_excel(data or SAMPLE_DATA)
        return buf.getvalue()

    def test_parse_plug_details(self):
        """Parse a generated workbook's plugs and verify data matches."""
        result = parse_wbd_excel(self._generate_bytes())
        plugs = result["plugs"]

        assert len(plugs) == 2
        assert plugs[0]["plug_number"] == 1
        assert plugs[0]["type"] == "cement"
        assert plugs[0]["top_ft"] == 5000.0
        assert plugs[0]["bottom_ft"] == 5100.0
        assert plugs[0]["sacks"] == 50.0
        assert plugs[0]["cement_class"] == "H"

        assert plugs[1]["plug_number"] == 2
        assert plugs[1]["type"] == "bridge_plug"

    def test_parse_geometry_casing(self):
        """Parse casing strings from Sheet 3."""
        result = parse_wbd_excel(self._generate_bytes())
        casings = result["well_geometry"]["casing_strings"]

        assert len(casings) == 2
        # Find surface casing
        surface = next(c for c in casings if "surface" in c["string"].lower())
        assert surface["size_in"] == 8.625
        assert surface["bottom_ft"] == 500.0

    def test_parse_geometry_formations(self):
        """Parse formation tops from Sheet 3."""
        result = parse_wbd_excel(self._generate_bytes())
        formations = result["well_geometry"]["formation_tops"]

        assert len(formations) == 2
        austin = next(f for f in formations if "austin" in f["formation"].lower())
        assert austin["depth"] == 2000.0

    def test_parse_geometry_perforations(self):
        """Parse perforations from Sheet 3."""
        result = parse_wbd_excel(self._generate_bytes())
        perfs = result["well_geometry"]["perforations"]

        assert len(perfs) == 1
        assert perfs[0]["top_ft"] == 5500.0
        assert perfs[0]["bottom_ft"] == 5600.0

    def test_round_trip_equivalence(self):
        """Generate → parse → verify core data matches input."""
        excel_bytes = self._generate_bytes()
        result = parse_wbd_excel(excel_bytes)

        # Plugs match
        for orig, parsed in zip(SAMPLE_DATA["comparisons"], result["plugs"]):
            assert parsed["plug_number"] == orig["plug_number"]
            assert parsed["type"] == orig["actual_type"]
            assert parsed["top_ft"] == orig["actual_top_ft"]
            assert parsed["bottom_ft"] == orig["actual_bottom_ft"]

        # Casings match
        orig_casings = SAMPLE_DATA["well_geometry"]["casing_strings"]
        parsed_casings = result["well_geometry"]["casing_strings"]
        assert len(parsed_casings) == len(orig_casings)
        for orig, parsed in zip(orig_casings, parsed_casings):
            assert parsed["bottom_ft"] == orig["bottom_ft"]

    def test_parse_empty_plugs(self):
        """Parse workbook with no plugs — should return empty list."""
        result = parse_wbd_excel(self._generate_bytes(EMPTY_GEOMETRY_DATA))
        assert result["plugs"] == []

    def test_parse_invalid_file(self):
        """Non-Excel bytes should raise ValueError."""
        with pytest.raises(ValueError, match="not a valid Excel"):
            parse_wbd_excel(b"this is not excel data")

    def test_parse_missing_sheet(self):
        """Workbook without required sheets should raise ValueError."""
        wb = openpyxl.Workbook()
        wb.active.title = "Random Sheet"
        buf = BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="Plug Details"):
            parse_wbd_excel(buf.getvalue())

    def test_parse_deep_well(self):
        """Round-trip for a deep well with tubing and tools."""
        result = parse_wbd_excel(self._generate_bytes(DEEP_WELL_DATA))

        assert len(result["plugs"]) == 1
        assert result["plugs"][0]["top_ft"] == 14000.0

        casings = result["well_geometry"]["casing_strings"]
        assert len(casings) == 3

        tubing = result["well_geometry"]["tubing"]
        assert len(tubing) == 1
        assert tubing[0]["size_in"] == 2.875

        tools = result["well_geometry"]["tools"]
        assert len(tools) == 1

    def test_parse_modified_plug_depth(self):
        """Generate, modify a plug depth in the Excel, parse, verify new depth."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        wb = openpyxl.load_workbook(buf)
        ws = wb["Plug Details"]

        # Modify plug 1's top depth from 5000 to 4900
        ws.cell(row=3, column=3, value=4900.0)

        modified_buf = BytesIO()
        wb.save(modified_buf)

        result = parse_wbd_excel(modified_buf.getvalue())
        assert result["plugs"][0]["top_ft"] == 4900.0  # Modified value
        assert result["plugs"][0]["bottom_ft"] == 5100.0  # Unchanged

    def test_parse_named_range_fallback(self):
        """Remove named ranges, verify parser falls back to header scanning."""
        buf = generate_wbd_excel(SAMPLE_DATA)
        wb = openpyxl.load_workbook(buf)

        # Remove all defined names
        for name in list(wb.defined_names.keys()):
            del wb.defined_names[name]

        modified_buf = BytesIO()
        wb.save(modified_buf)

        result = parse_wbd_excel(modified_buf.getvalue())
        # Should still parse successfully via fallback
        assert len(result["plugs"]) == 2
        assert len(result["well_geometry"]["casing_strings"]) == 2


# ---------------------------------------------------------------------------
# Additional edge case tests
# ---------------------------------------------------------------------------

class TestWbdExcelEdgeCases:
    """Additional edge case tests."""

    def test_none_values_in_comparisons(self):
        """Comparisons with None values for optional fields should work."""
        data = {
            **SAMPLE_DATA,
            "comparisons": [
                {
                    "plug_number": 1,
                    "actual_type": "cement",
                    "actual_top_ft": 1000.0,
                    "actual_bottom_ft": 1100.0,
                    "actual_sacks": None,
                    "actual_cement_class": None,
                    "actual_tagged_depth_ft": None,
                    "actual_placement_method": None,
                    "actual_woc_hours": None,
                    "actual_woc_tagged": None,
                },
            ],
        }
        buf = generate_wbd_excel(data)
        result = parse_wbd_excel(buf.getvalue())
        assert len(result["plugs"]) == 1
        assert result["plugs"][0]["sacks"] is None
        assert result["plugs"][0]["woc_tagged"] is None

    def test_multiple_formations(self):
        """Well with many formations should render correctly."""
        data = {
            **SAMPLE_DATA,
            "well_geometry": {
                **SAMPLE_DATA["well_geometry"],
                "formation_tops": [
                    {"formation": "Austin Chalk", "depth": 1000.0, "base_depth": 2000.0},
                    {"formation": "Eagle Ford", "depth": 2000.0, "base_depth": 3000.0},
                    {"formation": "Buda", "depth": 3000.0, "base_depth": 3500.0},
                    {"formation": "Georgetown", "depth": 3500.0, "base_depth": 4000.0},
                    {"formation": "Del Rio", "depth": 4000.0, "base_depth": 4500.0},
                    {"formation": "Edwards", "depth": 4500.0, "base_depth": 5000.0},
                ],
            },
        }
        buf = generate_wbd_excel(data)
        result = parse_wbd_excel(buf.getvalue())
        assert len(result["well_geometry"]["formation_tops"]) == 6
